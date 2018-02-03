# coding:utf-8
import sys
import os
from selenium import webdriver
from selenium.webdriver.common.keys import Keys
from time import sleep
from selenium.webdriver.support.ui import Select
from selenium.webdriver.common.action_chains import ActionChains
import openpyxl as px
import pywinauto
import configparser

#ユーザー設定
inifile = configparser.ConfigParser()
inifile.read('.\\config.ini', 'UTF-8')

username = inifile.get('shop_settings', 'username')
password = inifile.get('shop_settings', 'password')
shop_id = inifile.get('shop_settings', 'shop_id')
start_day = inifile.get('crawl_settings', 'start_day')
day = int(inifile.get('crawl_settings', 'day'))
printout = inifile.get('crawl_settings', 'printout')

#保存するExcelファイルの選択と有効化
wb = px.load_workbook('work_data.xlsx')
ws = wb.active

#chromedriverの指定
driver = webdriver.Chrome(executable_path='.\\chromedriver.exe')

def login(username,password):
    # 管理画面アクセス
    driver.get("https://touchontime.com/admin")
    #最大化
    driver.maximize_window()
    # ログイン処理
    driver.find_element_by_id("login_id").send_keys(username)
    driver.find_element_by_id("login_password").send_keys(password)
    driver.find_element_by_id("login_password").send_keys(Keys.ENTER)

def skip():
    #適当な座標をクリックしてチュートリアル画面スキップ
    actions = ActionChains(driver)
    actions.move_by_offset(100,100)
    actions.click()
    actions.perform()

def date_setting(start_day,shop_id):
    #日別集計画面選択
    driver.find_element_by_xpath("//*[@id='daily_working_link']").click()
    sleep(2)
    #店舗の設定
    shop_element = driver.find_element_by_name('selected_section_id')
    shop_select_element = Select(shop_element)
    shop_select_element.select_by_value(shop_id)

    #日付の設定
    for i in range(10):
        driver.find_element_by_id("parts_daily_select_picker").send_keys(Keys.BACKSPACE)
    driver.find_element_by_id("parts_daily_select_picker").send_keys(start_day)
    driver.find_element_by_id("parts_daily_select_picker").send_keys(Keys.ENTER)
    driver.find_element_by_id("display_button").click()
    sleep(2)

def page_print():
    #印刷ダイアログを開く
    pywinauto.keyboard.SendKeys("+^P")
    sleep(1)
    a_check = lambda:pywinauto.findwindows.find_windows(title=u'印刷', class_name='#32770')[0]
    dialog = pywinauto.timings.WaitUntilPasses(5, 1, a_check)
    pwa_app = pywinauto.Application()
    pwa_app.connect(handle=dialog)
    #Windowsの印刷設定画面
    window1 = pwa_app[u'印刷']
    window1.Wait('ready')
    sleep(1)
    button1 = window1[u'詳細設定(&R)']
    button1.Click()
    #Brotherの印刷設定画面
    window2 = pwa_app[u'印刷設定']
    window2.Wait('ready')
    window2.TabControl.Select(0)
    sleep(1)
    window2[u'横(&E)'].Click()
    #印刷
    sleep(1)
    window2[u'OK'].Click()
    window1[u'印刷(&P)'].Click()

def crawl():
    i=1
    while i<=day:
        #日付の取得
        element = driver.find_element_by_xpath("//*[@id='parts_daily_select_picker']")
        a = element.get_attribute("value")
        ws['%s%d' % ("A",i+1)].value = a

        #所定時間の取得
        element = driver.find_element_by_xpath("//*[@id='tab-1']/div/div[2]/div[1]/table/tfoot/tr/td[13]/p")
        ws['%s%d' % ("B",i+1)].value = element.text

        #深夜時間の取得
        element = driver.find_element_by_xpath("//*[@id='tab-1']/div/div[2]/div[1]/table/tfoot/tr/td[16]/p")
        ws['%s%d' % ("C",i+1)].value = element.text

        #深夜残業時間の取得
        element = driver.find_element_by_xpath("//*[@id='tab-1']/div/div[2]/div[1]/table/tfoot/tr/td[17]/p")
        ws['%s%d' % ("D",i+1)].value = element.text

        #印刷する
        page_print() if printout == "on" else None

        #次のページに遷移
        driver.find_element_by_xpath("//*[@id='button_next_day']").click()
        sleep(1)
        i+=1

if __name__ == '__main__':
    login(username,password)
    skip()
    date_setting(start_day,shop_id)
    crawl()
    #保存
    wb.save('work_data.xlsx')
    #終了
    print("3秒後に終了します")
    sleep(3)
    driver.quit()
